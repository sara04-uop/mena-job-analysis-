"""
MENA Job Market Scraper — Graduation Project
=============================================
Scrapes job listings from Bayt.com across 4 MENA countries
(Jordan, UAE, Saudi Arabia, Egypt) for 8 academic majors/sectors.

What this script collects (one row per job):
    Job Title | Company | Location | Sector | Source |
    Experience Level | Date Posted | URL | Job Description

"""





import time
import random
import re
import logging
from datetime import datetime, timedelta
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, WebDriverException
)

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── LOGGING ───────────────────────────────────────────────────
# Logs appear in your terminal AND are saved to scraper_log.txt
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("scraper_log.txt", encoding="utf-8"),
    ]
)
log = logging.getLogger(__name__)


# ── CONFIG ────────────────────────────────────────────────────
MAX_PAGES        = 10     # how many result pages to scrape per keyword
MIN_DELAY        = 1.5    # minimum seconds to wait between page loads
MAX_DELAY        = 3.0    # maximum seconds to wait between page loads
OUTPUT_FILE      = "jordan_jobs_2000_sample (1).xlsx"  # output Excel file name
CHECKPOINT_EVERY = 100    # save a backup file every N description scrapes


# ── OUTPUT COLUMNS ─────────────────────────────────────────────
COLUMNS = [
    "Job Title",
    "Company",
    "Location",
    "Sector",
    "Source",
    "Experience Level",
    "Date Posted",
    "URL",
    "Job Description",
]

# Excel column widths (characters) for each column above
COL_WIDTHS = [35, 28, 22, 30, 14, 16, 15, 55, 90]


# ── COUNTRIES TO SCRAPE ───────────────────────────────────────
BAYT_COUNTRIES = [
    ("jordan",               "Jordan"),
    ("united-arab-emirates", "UAE"),
    ("saudi-arabia",         "Saudi Arabia"),
    ("egypt",                "Egypt"),
]


# ── SECTORS AND SEARCH KEYWORDS ───────────────────────────────
# For each sector we search these job-title keywords on Bayt.com.
MAJORS = {
    "Data Science": [
        "data scientist", "data analyst", "machine learning", "AI engineer",
        "data engineer", "NLP engineer", "deep learning", "analytics engineer", "big data",
    ],
    "Computer Science": [
        "software developer", "backend developer", "frontend developer",
        "full stack developer", "web developer", "mobile developer",
        "flutter developer", "react developer", "node developer", "php developer",
    ],
    "Software Engineering": [
        "software engineer", "DevOps engineer", "QA engineer", "cloud engineer",
        "solutions architect", "systems engineer", "test automation engineer",
        "site reliability engineer", "platform engineer",
    ],
    "Cyber Security": [
        "cybersecurity", "information security", "penetration tester", "SOC analyst",
        "network security engineer", "security analyst", "security engineer", "ethical hacker",
    ],
    "Business Intelligence": [
        "business intelligence", "BI developer", "power BI developer",
        "tableau developer", "data warehouse engineer", "reporting analyst",
        "BI analyst", "MIS analyst", "ETL developer",
    ],
    "E-Marketing & Digital Marketing": [
        "digital marketing", "SEO specialist", "social media manager",
        "content marketing", "Google Ads specialist", "performance marketing",
        "e-commerce manager", "digital marketing manager",
        "content creator", "email marketing specialist",
    ],
    "Business Administration": [
        "business analyst", "project manager", "operations manager", "HR manager",
        "product manager", "account manager", "general manager",
        "business development", "supply chain manager",
        "procurement manager", "administrative manager",
    ],
    "Accounting": [
        "accountant", "financial analyst", "auditor", "finance manager",
        "tax accountant", "cost accountant", "financial controller",
        "accounts payable", "accounts receivable",
        "payroll accountant", "budget analyst",
    ],
}



# ── HELPER FUNCTIONS ──────────────────────────────────────────

def pause(lo=None, hi=None):
    """
    Waits a random number of seconds between requests.
    Randomness makes the scraper look less like a bot.
    """
    time.sleep(random.uniform(lo or MIN_DELAY, hi or MAX_DELAY))


def clean_date(text):
    """
    Converts raw date text into a clean YYYY-MM-DD string.

    Job sites show dates in many formats:
      "3 days ago"    → today minus 3 days
      "2 weeks ago"   → today minus 14 days
      "2026-04-14"    → used as-is
      "April 14 2026" → parsed with format patterns
    Falls back to today's date if nothing matches.
    """
    if not text or not text.strip():
        return str(datetime.today().date())
    t = text.strip().lower()

    m = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    if m:
        return m.group(1)
    if any(x in t for x in ["just now", "hour", "minute", "today", "less than"]):
        return str(datetime.today().date())
    m = re.search(r"(\d+)\s*day", t)
    if m:
        return str((datetime.today() - timedelta(days=int(m.group(1)))).date())
    m = re.search(r"(\d+)\s*week", t)
    if m:
        return str((datetime.today() - timedelta(weeks=int(m.group(1)))).date())
    m = re.search(r"(\d+)\s*month", t)
    if m:
        return str((datetime.today() - timedelta(days=int(m.group(1)) * 30)).date())
    for fmt in ("%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y",
                "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return str(datetime.strptime(text.strip(), fmt).date())
        except:
            pass
    return str(datetime.today().date())


def infer_exp(title):
    """
    Infers the experience level from keywords in the job title.

    Platforms don't always state experience level explicitly,
    so we detect it from words like 'senior', 'junior', 'manager'.
    Defaults to 'Mid-level' when no clear signal is found.
    """
    t = title.lower()
    if any(k in t for k in ["senior", "sr.", "lead", "principal", "head of", "staff"]):
        return "Senior"
    if any(k in t for k in ["junior", "jr.", "entry", "graduate", "intern", "fresh"]):
        return "Junior"
    if any(k in t for k in ["manager", "director", "vp ", "chief", "executive", "ceo", "cto"]):
        return "Managerial"
    return "Mid-level"


def get_text(element, css, default=""):
    """Safely reads text from a CSS selector inside a card element."""
    try:
        return element.find_element(By.CSS_SELECTOR, css).text.strip() or default
    except:
        return default


def scroll_page(driver, times=3, wait=1.0):
    """
    Scrolls down the page to load lazy content.
    Some job sites only render cards when you scroll past them.
    """
    for _ in range(times):
        driver.execute_script("window.scrollBy(0, 700);")
        time.sleep(wait)


def make_record(title, company, loc, sector, source, date_raw, link):
    """
    Builds one job record as a dictionary.
    Job Description is left empty here — it gets filled in Phase 2.
    """
    return {
        "Job Title":        title.strip(),
        "Company":          company.strip() if company else "N/A",
        "Location":         loc,
        "Sector":           sector,
        "Source":           source,
        "Experience Level": infer_exp(title),
        "Date Posted":      clean_date(date_raw),
        "URL":              link,
        "Job Description":  "",
    }


# ── BROWSER SETUP ─────────────────────────────────────────────

def build_driver():
    """
    Launches an Edge browser configured to look like a real user.
    Hides the automation flag so websites don't detect and block the scraper.
    """
    opts = Options()
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--lang=en-US,en")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/147.0.0.0 Safari/537.36 Edg/147.0.3912.72"
    )
    local = Path(__file__).parent / "msedgedriver.exe"
    svc = Service(str(local)) if local.exists() else Service()
    driver = webdriver.Edge(service=svc, options=opts)
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"}
    )
    driver.set_page_load_timeout(30)
    return driver


# ── PHASE 1: SCRAPE LISTINGS ──────────────────────────────────

def scrape_bayt(driver, keyword, sector):
    """
    Searches Bayt.com for one keyword across all 4 MENA countries.
    Visits up to MAX_PAGES result pages and collects job listing cards.
    Does NOT visit individual job pages yet — just the search results.
    """
    records = []
    slug = re.sub(r"[^a-z0-9]+", " ", keyword.lower()).strip().replace(" ", "-")

    for country_slug, country_name in BAYT_COUNTRIES:
        base_url = f"https://www.bayt.com/en/{country_slug}/jobs/{slug}-jobs/"
        log.info(f"  [Bayt/{country_name}] Searching: '{keyword}'")

        for page in range(1, MAX_PAGES + 1):
            url = base_url if page == 1 else f"{base_url}?page={page}"
            try:
                driver.get(url)
                pause()
                WebDriverWait(driver, 12).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
            except:
                break

            # Each job listing on Bayt is inside a <li data-job-id="..."> tag
            cards = driver.find_elements(By.CSS_SELECTOR, "li[data-job-id]")
            if not cards:
                log.info(f"    No cards on page {page} — stopping")
                break
            log.info(f"    Page {page}: {len(cards)} jobs")

            for card in cards:
                try:
                    a_tags = card.find_elements(By.CSS_SELECTOR, "h2 a")
                    if not a_tags:
                        continue
                    title = a_tags[0].text.strip()
                    if not title:
                        continue
                    href = a_tags[0].get_attribute("href") or ""
                    link = href if href.startswith("http") else "https://www.bayt.com" + href

                    company  = get_text(card, ".t-default", "N/A")
                    loc      = country_name
                    date_raw = ""

                    for mi in card.find_elements(By.CSS_SELECTOR, "li.jb-meta-item"):
                        txt = mi.text.strip()
                        if not txt:
                            continue
                        if re.search(r"ago|day|week|month|\d{4}", txt.lower()) and not date_raw:
                            date_raw = txt
                        elif any(city in txt.lower() for city in
                                 ["amman", "dubai", "riyadh", "cairo", "jeddah", "abu dhabi"]):
                            loc = txt

                    if not date_raw:
                        for m_el in card.find_elements(By.CSS_SELECTOR, ".t-mute"):
                            txt = m_el.text.strip()
                            if re.search(r"ago|day|week|month|\d{4}", txt.lower()):
                                date_raw = txt
                                break

                    records.append(make_record(title, company, loc, sector, "Bayt", date_raw, link))

                except Exception as e:
                    log.debug(f"    Card error: {e}")

            if len(cards) < 8:
                break

    log.info(f"  [Bayt] '{keyword}' → {len(records)} collected")
    return records


def run_phase1(driver):
    """
    Phase 1: Iterates over all sectors and keywords, runs the Bayt scraper,
    removes duplicate URLs, and returns a clean DataFrame of job listings.
    """
    all_records = []

    for sector, keywords in MAJORS.items():
        log.info(f"\n{'='*60}\n  SECTOR: {sector}\n{'='*60}")
        for keyword in keywords:
            try:
                recs = scrape_bayt(driver, keyword, sector)
                all_records.extend(recs)
                log.info(f"  ✓ Bayt/'{keyword}': +{len(recs)} | total={len(all_records)}")
            except WebDriverException as e:
                log.error(f"  ✗ Bayt/'{keyword}': {str(e)[:100]}")
            except Exception as e:
                log.error(f"  ✗ Bayt/'{keyword}': {str(e)[:100]}")

    df = pd.DataFrame(all_records, columns=COLUMNS)
    before = len(df)
    df.drop_duplicates(subset=["URL"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    log.info(f"\n  Phase 1 complete.")
    log.info(f"  Raw records: {before} → After dedup: {len(df)}")
    log.info(f"    Bayt: {len(df)}")
    return df


# ── PHASE 2: SCRAPE JOB DESCRIPTIONS ─────────────────────────

def scrape_one_description(driver, url):
    """
    Opens a single job posting page and extracts the full description text.

    Tries multiple CSS selectors because Bayt's layout differs per job.
    Requires at least 80 characters to avoid picking up navigation text.
    Returns "Not Found", "Timeout", or "Error" if extraction fails.
    """
    try:
        driver.get(url)
        time.sleep(random.uniform(3.0, 5.0))
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
        driver.execute_script("window.scrollBy(0, 500);")
        time.sleep(1.5)

        selectors = [
            ".card-content.p20t.is-spaced",
            ".card-content.is-spaced",
            ".card-content",
            ".t-break",
            "#job_card .card-content",
            "section.card .card-content",
        ]
        for selector in selectors:
            try:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for el in elements:
                    text = el.text.strip()
                    if text and len(text) > 80:
                        return text
            except:
                continue

        try:
            card = driver.find_element(By.CSS_SELECTOR, "#job_card")
            text = card.text.strip()
            if len(text) > 80:
                return text
        except:
            pass

        return "Not Found"

    except TimeoutException:
        log.warning(f"    Timeout: {url}")
        return "Timeout"
    except Exception as e:
        log.warning(f"    Error: {url} — {str(e)[:60]}")
        return "Error"


def run_phase2(driver, df):
    """
    Phase 2: Visits every job URL collected in Phase 1 and scrapes
    the full job description text into the Job Description column.

    Saves a checkpoint file every CHECKPOINT_EVERY rows to protect
    your data in case the script is interrupted.
    """
    log.info(f"\n{'='*60}")
    log.info(f"  Phase 2: Scraping descriptions for {len(df)} jobs")
    log.info(f"  Estimated time: ~{len(df) // 12} minutes")
    log.info(f"{'='*60}")

    descriptions = [""] * len(df)
    success = 0
    failed  = 0

    for i, (idx, row) in enumerate(df.iterrows(), 1):
        url = str(row.get("URL", ""))

        if not url.startswith("http"):
            descriptions[i - 1] = "No URL"
            failed += 1
            continue

        log.info(f"  [{i}/{len(df)}] {url[:75]}")
        desc = scrape_one_description(driver, url)
        descriptions[i - 1] = desc

        if desc not in ["Not Found", "Timeout", "Error", "No URL"]:
            success += 1
            log.info(f"    ✓ {len(desc)} chars")
        else:
            failed += 1
            log.info(f"    ✗ {desc}")

        # Checkpoint save every N rows
        if i % CHECKPOINT_EVERY == 0:
            df_temp = df.copy()
            df_temp["Job Description"] = descriptions
            df_temp.to_excel("jordan_jobs_checkpoint.xlsx", index=False, sheet_name="All Jobs")
            log.info(f"    ✅ Checkpoint saved — Success: {success} | Failed: {failed}")

    df["Job Description"] = descriptions
    log.info(f"\n  Phase 2 complete. Success: {success} | Failed: {failed}")
    return df


# ── PHASE 3: SAVE TO EXCEL ────────────────────────────────────

HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN      = Side(style="thin", color="CCCCCC")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_header(ws):
    """Writes the styled header row to a worksheet."""
    for ci, (name, width) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        c = ws.cell(1, ci, name)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = HDR_ALIGN
        c.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def write_data_row(ws, row_num, row_dict):
    """Writes one job record as a styled row."""
    for ci, col in enumerate(COLUMNS, 1):
        val = row_dict.get(col, "")
        c = ws.cell(row_num, ci, val)
        c.font      = Font(name="Arial", size=9)
        c.border    = BORDER
        c.alignment = Alignment(
            vertical="center",
            wrap_text=(col in ("Job Title", "Job Description"))
        )
        if col == "URL" and val and str(val).startswith("http"):
            c.hyperlink = str(val)
            c.font = Font(name="Arial", size=9, color="0563C1", underline="single")


def save_to_excel(df, filepath):
    """
    Saves the full dataset to a single formatted Excel sheet.
    The sheet contains every job record with a styled header,
    column filters, and frozen top row.
    """
    log.info(f"\n  Saving {len(df)} records to {filepath}...")
    wb = Workbook()

    # ── All Jobs Sheet ─────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Jobs"
    write_header(ws_all)
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        write_data_row(ws_all, ri, row.to_dict())
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(df) + 1}"

    wb.save(filepath)
    log.info(f"  ✅ Saved: {filepath}  ({len(df)} rows)")


# ── MAIN ──────────────────────────────────────────────────────

if __name__ == "__main__":
    log.info("=" * 60)
    log.info("  MENA Job Market Scraper — Graduation Project")
    log.info(f"  Sectors  : {len(MAJORS)}")
    log.info(f"  Keywords : {sum(len(v) for v in MAJORS.values())}")
    log.info(f"  Countries: Jordan, UAE, Saudi Arabia, Egypt")
    log.info(f"  Source   : Bayt.com")
    log.info("=" * 60)

    driver = build_driver()

    try:
        # Phase 1: collect all job listings from search pages
        log.info("\n>>> PHASE 1: Scraping job listings...")
        df = run_phase1(driver)

        if df.empty:
            log.error("No records collected. Check scraper_log.txt.")

        else:
            # Phase 2: visit each URL and get the full description
            log.info("\n>>> PHASE 2: Scraping job descriptions...")
            df = run_phase2(driver, df)

            # Phase 3: write everything to a formatted Excel file
            log.info("\n>>> PHASE 3: Saving to Excel...")
            save_to_excel(df, OUTPUT_FILE)

            log.info("\n" + "=" * 60)
            log.info(f"  ALL DONE!")
            log.info(f"  Output : {OUTPUT_FILE}")
            log.info(f"  Total  : {len(df)} jobs")
            log.info("=" * 60)

    finally:
        driver.quit()  # always close the browser, even if something crashes